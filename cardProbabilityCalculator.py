import openpyxl, os, sys

if __name__ == "__main__":
    while True:
        
        wb = openpyxl.load_workbook(filename='Card_Probabilities.xlsx')
        faceSheet = wb['faceCardProbabilities']
        otherSheet = wb['otherCardProbabilities']

        for col in faceSheet.iter_cols(min_row=2, min_col=2, max_col=42, max_row=16):
            for cell in col:
                try:drawnCardProb = faceSheet[f'A{cell.row}'].value/(faceSheet[f'{cell.column_letter}1'].value + faceSheet[f'A{cell.row}'].value)
                except ZeroDivisionError:drawnCardProb=0
                print(drawnCardProb)
                cell.value = drawnCardProb

        for col in otherSheet.iter_cols(min_row=2, min_col=2, max_col=42, max_row=16):
            for cell in col:
                try:drawnCardProb = otherSheet[f'{cell.column_letter}1'].value/(otherSheet[f'{cell.column_letter}1'].value + otherSheet[f'A{cell.row}'].value)
                except ZeroDivisionError:drawnCardProb=0
                print(drawnCardProb)
                cell.value = drawnCardProb
        
        wb.save('Card_Probabilities.xlsx')

        sys.exit()