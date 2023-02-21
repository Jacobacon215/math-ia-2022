import openpyxl, num2words, os, sys, math, itertools, gc

class grab_workbook:
    def __init__(self, filename:str) -> None:
        self.filename=filename
        pass
    def __enter__(self):
        pass

    pass


class possible_number_of_face_cards: # this is xn and pn bundled together
    def __init__(self, how_many:int, total_number_of_cards:int) -> None:
        # how many is the possible number of face cards
        self.how_many = how_many
        if self.how_many > 16:
            raise ValueError('Maximum exceeded: more than 16 cards drawn')
        elif self.how_many < 1:
            raise ValueError('Minimum exceeded: fewer than 1 card drawn')

        # how many total cards are there
        self.total_number_of_cards = total_number_of_cards
        if self.total_number_of_cards > 16:
            raise ValueError('Maximum exceeded: more than 16 cards drawn')
        elif self.total_number_of_cards < 1:
            raise ValueError('Minimum exceeded: fewer than 1 card drawn')
        

        # logic to determine probability
        workbook,deck,probability,permutation_probability,list_of_total_cards_and_types,alphabet_dict = openpyxl.load_workbook(filename='Card_Probabilities.xlsx'),54,0,1,[],{40:'B',39:'C',38:'D',37:'E',36:'F',35:'G',34:'H',33:'I',32:'J',31:'K',30:'L',29:'M',28:'N',27:'O',26:'P',25:'Q',24:'R',23:'S',22:'T',21:'U',20:'V',19:'W',18:'X',17:'Y',16:'Z',15:'AA',14:'AB',13:'AC',12:'AD',11:'AE',10:'AF',9:'AG',8:'AH',7:'AI',6:'AJ',5:'AK',4:'AL',3:'AM',2:'AN',1:'AO',0:'AP'}
        faceSheet = workbook['faceCardProbabilities']
        otherSheet = workbook['otherCardProbabilities']

        numbersOfCardsDict = dict()
        numbersOfCardsDict['numFace'] = 14
        numbersOfCardsDict['numOther'] = 40

        # a) generate list with how_many * 'face' and total_number_of_cards - how_many * 'other'
        # b) generate a list of permutations using that list
        # c) for each permutation tuple, draw cards in that order and multiply the probability of drawing that card to get the probability of that permutation occuring
        # d) add the probabilities of all the permutations together to get the probability of drawing how_many face cards when total_number_of_cards cards are drawn

        # a)
        for i in range(self.how_many):
            list_of_total_cards_and_types.append('face')
        for i in range(self.total_number_of_cards - self.how_many):
            list_of_total_cards_and_types.append('other')
        
        # b)
        def perms(num_a, num_b): # this is all ripped directly from the stackexchange answer to my question at https://stackoverflow.com/questions/75415051/how-to-check-for-repeated-yield-from-a-python-generator-function?noredirect=1#comment133067580_75415051
            res = []
            # A bit more efficient if we choose the indexes of the "minority" element
            if num_a < num_b:
                elem_sm, elem_lg = 'face', 'other'
                num_sm = num_a
            else:
                elem_sm, elem_lg = 'other', 'face'
                num_sm = num_b
            base_perm = [elem_lg] * (num_a + num_b)
            
            for c in itertools.combinations(range(num_a + num_b), num_sm):
                # c is the tuple of indexes that will be occupied by
                # the "minority" element
                perm = base_perm[:]
                for i in c:
                    perm[i] = elem_sm
                res.append(perm)

            print(res) # mine :)
            return res
        


        # c)
        for permutation in perms(num_a=self.how_many, num_b=self.total_number_of_cards-self.how_many):
            print(permutation)
            for card in permutation:
                
                if numbersOfCardsDict['numFace'] < 1:
                    permutation_probability = 0
                    break


                card_probability = faceSheet[f'{alphabet_dict[numbersOfCardsDict["numOther"]]}{16-numbersOfCardsDict["numFace"]}'].value if card == 'face' else otherSheet[f'{alphabet_dict[numbersOfCardsDict["numOther"]]}{16-numbersOfCardsDict["numFace"]}'].value
                permutation_probability *= card_probability

                deck -= 1
                numbersOfCardsDict[f'num{card.capitalize()}'] -= 1
               # print(numbersOfCardsDict[f'num{card.capitalize()}'])
            
            # d)
            probability += permutation_probability

            permutation_probability = 1

            numbersOfCardsDict['numFace'], numbersOfCardsDict['numOther'] = 14, 40
            
        self.probability = probability
    
    def __str__(self) -> str:
        return f'The probability of drawing {self.how_many} face cards when drawing {self.total_number_of_cards} total cards is {self.probability}'


if __name__ == "__main__":
    while True:

        print(possible_number_of_face_cards(how_many=1, total_number_of_cards=2))

        
        # this part calculates the probabilities
        workbook = openpyxl.load_workbook(filename='Draw_Number_Probabilities.xlsx')
        worksheet = workbook['Face_Sheet']

        possible_numbers_of_face_cards_dict = dict()

        for col in worksheet.iter_cols(min_row=2, min_col=2, max_col=17, max_row=17):
            for cell in col:
                if cell.value != 'N/A':
                    possible_numbers_of_face_cards_dict[f'{cell.column_letter}{cell.row}'] = possible_number_of_face_cards(how_many=worksheet[f"A{cell.row}"].value, total_number_of_cards=worksheet[f"{cell.column_letter}1"].value)
                    cell.value = possible_numbers_of_face_cards_dict[f'{cell.column_letter}{cell.row}'].probability
                    pass
                pass
            pass

        workbook.save('Draw_Number_Probabilities.xlsx')

        # this part calculates the average number of cards for a given number of cards drawn
        workbook = openpyxl.load_workbook(filename='Draw_Number_Probabilities.xlsx')
        worksheet = workbook['Face_Sheet']

        averageDict = dict()
        for i in range(1,17):
            averageDict[f'{num2words.num2words(i)}Average'] = 0


        for col in worksheet.iter_cols(min_row=2, min_col=2, max_col=17, max_row=17):
            for cell in col:
                if cell.value != 'N/A':
                    averageDict[f'{num2words.num2words(worksheet[f"{cell.column_letter}1"].value)}Average'] += (possible_numbers_of_face_cards_dict[f'{cell.column_letter}{cell.row}'].probability * possible_numbers_of_face_cards_dict[f'{cell.column_letter}{cell.row}'].how_many)
                    worksheet[f'{cell.column_letter}19'].value = averageDict[f'{num2words.num2words(worksheet[f"{cell.column_letter}1"].value)}Average']
                    pass
                pass
            pass

        workbook.save('Draw_Number_Probabilities.xlsx')



        sys.exit()

## TODO:
#  - use probabilities from spreadsheet to calculate probability. like if
#    the deck has 3 face cards and 10 other cards just go to that cell on
#    the sheet and pull from the data instead of natively calculating the
#    probability.

